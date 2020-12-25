from datetime import datetime
import pandas as pd
import numpy as np
import xlrd
import openpyxl
from openpyxl import Workbook
from openpyxl.drawing import image
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet import page
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font
import xlsxwriter

#from.views import chemin_message_frais


def formater_message_frais(fichier):
    with open(fichier, 'r', encoding="utf8", errors='ignore') as file:  # encoding à retirer, le errors sert à ignorer le caractere ø en position 75, donc je l'ai enlevé dans la premiere fonction replace
        filedata = file.read()

    # suppression des données inutiles (respectez cet ordre chronologique pour les .replace)
    # on remplace cette ligne par le descrpitif qui nous intéresse
    filedata = filedata.replace(
        'N de commande client;Date de depart;Code article;Designation;Quantite;Lot;DLUO;Poids net;',  # <--virgule est ici
        '"Document commercial";"Date de départ";"Article";"Désignation";"Quantité commandée";"Lot";"DLUO";"Poids net";\n'
        )
    filedata = filedata.replace('*** FIN DE RAPPORT ***', '')
    filedata = filedata.replace(',', '.')  # car les poids sont en XX,XX kg je les veux en XX.XX kg
    filedata = filedata.replace(';;;;', ',')  # on enlève tous les multiples ;;;;; par des "une seule virgule"
    filedata = filedata.replace(';;;', ',')
    filedata = filedata.replace(';', ',')

    # Ecraser les données modifiées dans le fichier
    with open(fichier, 'w') as file:
        file.write(filedata)

    # Supprimer la ligne "TOTAL" entre deux codes produits (leurs lignes sont < à 21 caractères)
    ma_liste = []  # je crée une liste pour pouvoir ensuite travailler avec les méthodes propres aux listes (.append() etc)
    with open(fichier, 'r') as file:
        lines = file.readlines()  # cette méthode permet de mettre les données sous forme de liste
        for line in lines:
            if len(line) > 21:  # pour éliminer la ligne pré-TOTAL qui est + courte que les autres lignes ( < 21 )
                ma_liste.append(line)

    # On se débarasse des sauts de lignes "\n" entre deux groupes d'articles
    liste_temporaire = []
    for e in ma_liste:
        if e != '\n':  # ne concerne que les "\n" isolés, hors des lignes produits
            liste_temporaire.append(e)
    ma_liste = liste_temporaire

    # on définit un index, qui sera le point 0 du début de ma_liste
    index_debut = ma_liste.index(
        '"Document commercial","Date de départ","Article","Désignation","Quantité commandée","Lot","DLUO","Poids net",\n')
    ma_liste = ma_liste[index_debut:-1]
    # garde tout depuis cet index jusqu'à la fin du doc moins le dernier élément TOTAL GLOBAL
    # A partir de maintenant, on a juste une liste avec les infos qui nous intéresse et pas d'autres infos inutiles
    # A partir d'ici, on a des lignes qui commencent par une virugle sans le numéro de commande


    # ------------------------------------------------------------------------------------------------------
    # --------------- petite parenthèse : on cherche le numéro de la commande à partir d'ici ---------------
    # ------------------------------------------------------------------------------------------------------

    liste_num_cde = []  # on loop pour obtenir le numéro sous forme de liste
    for n in ma_liste[1]:  # l'indice[0] représente mes colonnes donc le num commande est ds la ligne suivante (0009456742)
        if n == ',':  # dès qu'on rencontrera une virgule, la loop s'arrête
            break
        else:
            liste_num_cde.append(n)
    # on obtient un résultat comme : liste_num_cde = ['0', '0', '0', '9', '4', '3', '7', '7', '1', '2']


    # on veut savoir combien il y a de 0 inutiles avant le numéro de commande, sans supprimer les futurs 0 éventuels
    compteur_de_zeros = 0  # va me servir de compteur de zéros
    for n in liste_num_cde:
        if n == '0':
            compteur_de_zeros += 1
        if n != '0':
            break  # dès que je tombe sur un chiffre différent de zéro, la loop s'arrête

    num_cde = ''  # je crée une empty string pour y ajouter les éléments de liste_num_cde sans les 0,
    # grâce à compteur_de_zeros je sais que je dois commencer à l'index compteur_de_zeros

    # liste_num_cde[compteur_de_zeros:] = je loop en dehors des zéros inutiles et j'ajoute à ma string num_cde
    for z in liste_num_cde[compteur_de_zeros:]:
        num_cde += z
    # donc ma liste_num_cde = ['0', '0', '0', '9', '4', '3', '7', '7', '1', '2'] est maitenant devenue num_cde = 9437712

    # ------------------------------------------------------------------------------------------------------
    # --------------- RETOUR AU FORMATAGE DU FICHIER MAINTENANT QU'ON CONNAIT LE num_cde -------------------
    # ------------------------------------------------------------------------------------------------------

    # retour à ma_liste, qui contient toujours les numéros de commande avec les 0009437712 (= liste_num_cde)
    num_cde_avec_zeros = ''  # j'ai besoin du num de commande avec les zéros pour l'ajouter là il n'y est pas
    # sinon j'ai une liste avec des num_cde sans 0 et avec des 0, je veux avoir tout au même format avant de traiter
    for n in liste_num_cde:
        num_cde_avec_zeros += n  # ici j'obtiens num_cde_avec_zeros = 0009437712, ensuite je l'ajoute à ma_liste :

    liste_temporaire = []  # on crée une liste temporaire
    for n in ma_liste[1:]:  # on commence par l'indice car car l'indice [0] correspond à mes colones
        if n[0] != ',':  # si l'élément ne commence pas par une virgule, on l'ajoute à la liste
            liste_temporaire.append(n)
        else:  # s'il commence par une virgule, on veut y ajouter le 0009437712
            n = str(num_cde_avec_zeros) + n
            liste_temporaire.append(n)

    colones = ma_liste[0]  # on ajoute les colones à la nouvelle liste temporaire
    liste_temporaire.insert(0, colones)
    ma_liste = liste_temporaire  # on écrase les données à ma_liste

    with open(fichier, 'w', encoding='utf-8') as file:
        for n in ma_liste:
            file.write(n)

    # ------------------------------------------------------------------------------------------------------
    # --------------- LE FORMATAGE DU FICHIER USINE EST FINI, ON VIENT DE CREER LE .CSV --------------------
    # ------------------------------------------------------------------------------------------------------


    data_message_frais = pd.read_csv(fichier, encoding='utf-8')
    df = pd.DataFrame(data_message_frais, columns=[
        "Date de départ", "Article", "Désignation", "Quantité commandée", "Lot", "DLUO", "Poids net"
    ])  # on crée le dataframe à partir du .txt et on désigne les colones qu'on veut garder

    df2 = df.groupby(
        ["Date de départ", "Article", "Désignation", "Lot", "DLUO"], as_index=False).sum()
    # ici on groupe par rapport à ces colones, car si le lot ou la date de sortie varie pour un même code produit, on veut
    # deux lignes bien distinctes.
    # On ajoute.sum() pour que les qtés commandées et les poids soient additionnés pour chaque groupage
    
    chemin_sanitaire_final = str(fichier[:-4] + '.xlsx')

    # Create a Pandas Excel writer using XlsxWriter as the engine.
    writer = pd.ExcelWriter(chemin_sanitaire_final, engine='xlsxwriter')
    # https://xlsxwriter.readthedocs.io/working_with_pandas.html
    # si besoin de modifier le format date

    # Convert the dataframe to an XlsxWriter Excel object.
    df2.to_excel(writer, sheet_name='Sheet1', index=None, startrow=12)

    # Close the Pandas Excel writer and output the Excel file.
    writer.save()

# ----------------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------

    # Maintenant on le rouvre avec Openpyxl pour y faire les modifs suivantes :
    wb = openpyxl.load_workbook(chemin_sanitaire_final)
    ws = wb["Sheet1"] # on définit sur quelle feuille on fait les modifs
    ws['B2'] = "Expéditeur/ Shipper:"
    ws['B2'].font = Font(italic=True)
    ws['B3'] = "Bel S.A."
    ws['B3'].font = Font(bold=True, color='00008000')
    ws['B4'] = "2 Allée de Longchamp"
    ws['B5'] = "92150 Suresnes"
    ws['B6'] = "Tél.: +33 (0)1 84 02 72 50"
    ws['B7'] = "Capital social 10308502,50 € "
    ws['B8'] = "SIREN 542088067 – RCN Nanterre – NAF 1051C"
    ws['D10'] = f"Order number : {num_cde}"
    ws['D10'].font = Font(bold=True, color='00008000')


    logo_bel = "/home/bel/bel.pythonanywhere.com/bel/mes-statics/images/logo-pour-excel.png"
    img = openpyxl.drawing.image.Image(logo_bel)  # on crée l'objet logo pour pouvoir l'ajouer au excel
    ws.add_image(img, anchor="A2")

    openpyxl.worksheet.worksheet.Worksheet.set_printer_settings(ws, paper_size=9, orientation='landscape')
    # paper_size=9 == format A4, voir documentation
    ws.sheet_properties.pageSetUpPr.fitToPage = True  # Ajuster à une page


    # redimensionne la largeur des colonnes
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 22
    ws.column_dimensions['G'].width = 22
    ws.column_dimensions['H'].width = 22
    ws.column_dimensions['I'].width = 22

    # centrer (mise en page) les valeurs du tableau :
    rows = range(13, 100)
    columns = [1, 2, 4, 5, 6, 7, 8]
    for row in rows:
        for col in columns:
            ws.cell(row, col).alignment = Alignment(horizontal='center')

    wb.save(chemin_sanitaire_final)

 